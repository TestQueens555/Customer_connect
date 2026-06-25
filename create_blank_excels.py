"""
create_blank_excels.py
Creates blank Login.xlsx and BugReport xlsx so Excel MCP can write to them.
Run: python create_blank_excels.py
"""
import subprocess, sys

# Install openpyxl if needed
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from datetime import date
import os

today = date.today().strftime("%d-%b-%Y")

base = r"D:\Claude\QA_Projects\CustomerConnect"

# 1. Login.xlsx — Featurewise Test Report
fw_path = os.path.join(base, "Featurewise Test Report", "Login.xlsx")
wb = Workbook()
wb.active.title = "Test Execution"
wb.create_sheet("Bug Report")
wb.save(fw_path)
print(f"Created: {fw_path}")

# 2. BugReport_DD-Mon-YYYY.xlsx — Daily Bug Report
db_path = os.path.join(base, "Daily Bug Report", f"BugReport_{today}.xlsx")
wb2 = Workbook()
wb2.active.title = "Daily Bug Report"
wb2.create_sheet("Test Summary")
wb2.save(db_path)
print(f"Created: {db_path}")

print("Done — blank Excel files created. Claude will now fill them via Excel MCP.")
